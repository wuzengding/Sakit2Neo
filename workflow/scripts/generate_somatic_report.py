#!/usr/bin/env python3
#generate_somatic_report.py
"""
Generate a comprehensive, annotated Excel report for somatic variants and neoantigen screening.

Key Updates in this version:
- Blacklist Interception: Rapidly removes TILs (IGHV/TRBV) and high-artifact genes (MUC/HLA/TTN).
- Whitelist Rescue: Forces inclusion of critical hotspots (e.g., BRAF V600E) regardless of coverage/VAF.
- Aware Flagging: Tags critical pathway mutations (APM Loss, Spliceosome, MMR) for clinical warning.
- Sample Summary Sheet: Creates a cover page with critical clinical warnings (e.g., B2M loss).
- Bugfix: Added robust handling for pysam NoneType returns in VCF INFO/FORMAT fields.
"""

import pandas as pd
import pysam
import logging
import argparse
import sys
import os
import re
import yaml
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# 1. 核心生物学知识库定义 (Blacklist / Whitelist / Aware)
# ==========================================

BLACKLIST_GENE_PREFIXES = ("IGH", "IGK", "IGL", "TRAV", "TRB", "TRG", "TRD", "HLA-", "MUC", "OR")
BLACKLIST_GENES_EXACT = {"TTN"}

def is_blacklisted(gene):
    if not isinstance(gene, str) or gene == "/": return False
    if gene in BLACKLIST_GENES_EXACT: return True
    if gene.startswith(BLACKLIST_GENE_PREFIXES): return True
    return False

WHITELIST_HOTSPOTS = {
    "BRAF": {"p.Val600Glu"},
    "KRAS": {"p.Gly12Cys", "p.Gly12Asp", "p.Gly12Val", "p.Gly13Asp", "p.Gln61His"},
    "NRAS": {"p.Gln61Arg", "p.Gln61Lys"},
    "PIK3CA": {"p.Glu545Lys", "p.His1047Arg"},
    "TP53": {"p.Arg175His", "p.Arg248Gln", "p.Arg273His"},
    "IDH1": {"p.Arg132His"}
}

def is_whitelisted(gene, hgvsp):
    if not isinstance(gene, str) or not isinstance(hgvsp, str): return False
    clean_hgvsp = hgvsp.split(':')[-1] if ':' in hgvsp else hgvsp
    return gene in WHITELIST_HOTSPOTS and clean_hgvsp in WHITELIST_HOTSPOTS[gene]

AWARE_MAP = {
    'B2M': 'APM_Loss (Immune Evasion)', 
    'TAP1': 'APM_Loss', 
    'TAP2': 'APM_Loss', 
    'NLRC5': 'APM_Loss', 
    'CIITA': 'APM_Loss', 
    'JAK1': 'IFNg_Pathway_Loss', 
    'JAK2': 'IFNg_Pathway_Loss',
    'SF3B1': 'Splice_Altered', 
    'U2AF1': 'Splice_Altered', 
    'SRSF2': 'Splice_Altered', 
    'ZRSR2': 'Splice_Altered',
    'MLH1': 'MMR_Defect (MSI-H)',
    'MSH2': 'MMR_Defect (MSI-H)', 
    'MSH6': 'MMR_Defect (MSI-H)', 
    'PMS2': 'MMR_Defect (MSI-H)', 
    'POLE': 'Polymerase_Defect',
    'POLD1': 'Polymerase_Defect'
}

# ==========================================
# 2. 基础配置与 VEP 权重
# ==========================================

CONSEQUENCE_SEVERITY_ORDER = [
    'transcript_ablation', 'splice_acceptor_variant', 'splice_donor_variant', 'stop_gained',
    'frameshift_variant', 'stop_lost', 'start_lost', 'transcript_amplification',
    'inframe_insertion', 'inframe_deletion', 'missense_variant', 'protein_altering_variant',
    'splice_donor_5th_base_variant', 'splice_region_variant', 'splice_donor_region_variant',
    'splice_polypyrimidine_tract_variant', 'incomplete_terminal_codon_variant', 'start_retained_variant',
    'stop_retained_variant', 'synonymous_variant', 'coding_sequence_variant', 'mature_miRNA_variant',
    '5_prime_UTR_variant', '3_prime_UTR_variant', 'non_coding_transcript_exon_variant', 'intron_variant',
    'NMD_transcript_variant', 'non_coding_transcript_variant', 'coding_transcript_variant',
    'upstream_gene_variant', 'downstream_gene_variant', 'TFBS_ablation', 'TFBS_amplification',
    'TF_binding_site_variant', 'regulatory_region_ablation', 'regulatory_region_amplification',
    'regulatory_region_variant', 'feature_elongation', 'feature_truncation', 'intergenic_variant', 'sequence_variant'
]

CONSEQUENCE_SEVERITY_MAP = {c: i for i, c in enumerate(CONSEQUENCE_SEVERITY_ORDER)}

PROTEIN_ALTERING_CONSEQUENCES = {
    'transcript_ablation', 'splice_acceptor_variant', 'splice_donor_variant', 'stop_gained',
    'frameshift_variant', 'stop_lost', 'start_lost', 'inframe_insertion', 'inframe_deletion',
    'missense_variant', 'protein_altering_variant'
}

def get_args():
    if 'snakemake' in globals():
        logging.info("Running in Snakemake mode.")
        return snakemake

    logging.info("Running in Standalone mode.")
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("--somatic-vcf", required=True)
    parser.add_argument("--varscan-vcf", required=False)
    parser.add_argument("--rna-counts-tumor", required=True)
    parser.add_argument("--cancer-genes", required=True)
    parser.add_argument("--gtf", required=True)
    parser.add_argument("--fasta", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--sample-id", required=True)
    parser.add_argument("--tiering-params-yaml", required=True)
    parser.add_argument("--log")
    args = parser.parse_args()

    class MockSnakemake:
        def __init__(self):
            self.input = {
                'somatic_vcf': args.somatic_vcf, 'varscan_vcf': args.varscan_vcf if args.varscan_vcf else [],
                'rna_counts_tumor': args.rna_counts_tumor, 'cancer_genes': args.cancer_genes,
                'gtf': args.gtf, 'fasta': args.fasta
            }
            self.output = {
                'xlsx_report_filter': args.output_xlsx.replace('.xlsx', '.Filtered_variants.xlsx'),
                'xlsx_report_Somatic': args.output_xlsx
            }
            self.params = {'sample_id': args.sample_id}
            self.log = [args.log] if args.log else None
            with open(args.tiering_params_yaml, 'r') as f:
                config_yaml = yaml.safe_load(f)
                self.params['somatic_tiering_params'] = config_yaml.get("somatic_tiering")
    return MockSnakemake()

def get_csq_header(vcf_reader):
    if "CSQ" not in vcf_reader.header.info: raise ValueError("CSQ field not found in VCF header.")
    match = re.search(r"Format: ([\w|]+)", vcf_reader.header.info["CSQ"].description)
    if match: return match.group(1).split("|")
    raise ValueError("Could not parse CSQ header format string.")

# [FIX] 使提取 INFO 字段更加安全，处理 (None,) 和 None 的情况
def get_safe_info_field(info, key, default):
    if key not in info:
        return default
    val = info.get(key)
    if isinstance(val, (list, tuple)):
        return val[0] if (len(val) > 0 and val[0] is not None) else default
    return val if val is not None else default

# [FIX] 使提取 FORMAT 字段更加安全
def get_safe_format_field(sample_format, key, default):
    val = sample_format.get(key)
    if val is None: 
        return default
    if isinstance(val, (list, tuple)):
        return val[0] if (len(val) > 0 and val[0] is not None) else default
    return val

def calculate_coding_ratio(csq_entries):
    if not csq_entries: return "0//0"
    coding = sum(1 for e in csq_entries if e.get("BIOTYPE") == "protein_coding")
    return f"{coding}//{len(csq_entries)}"

def calculate_AAchaged_ratio(csq_entries):
    if not csq_entries: return "0//0"
    altering = sum(1 for e in csq_entries if any(c in PROTEIN_ALTERING_CONSEQUENCES for c in e.get("Consequence", "").split('&')))
    return f"{altering}//{len(csq_entries)}"

def get_most_severe_consequence_rank(csq_entry):
    consequences = csq_entry.get("Consequence", "").split('&')
    min_rank = len(CONSEQUENCE_SEVERITY_ORDER) 
    for c in consequences:
        rank = CONSEQUENCE_SEVERITY_MAP.get(c, min_rank)
        if rank < min_rank: min_rank = rank
    return min_rank

def get_amino_acid_sub_status(consequences):
    if not any(c in PROTEIN_ALTERING_CONSEQUENCES for c in consequences): return "No_AA_Altered"
    if 'stop_gained' in consequences: return "Stop_Gained"
    if any(c in ['frameshift_variant', 'splice_acceptor_variant', 'splice_donor_variant', 'stop_lost'] for c in consequences):
        return "Long_AA_Altered"
    return "Single_AA_Altered"

def parse_vcf_record(record, alt_allele_index, csq_header, sample_id):
    alt_allele = record.alts[alt_allele_index]
    chrom, pos, ref = record.chrom, record.pos, record.ref
    variant_key = f"{chrom}-{pos}-{ref}-{alt_allele}"

    n_name = f"{sample_id}_normal" if f"{sample_id}_normal" in record.samples else "NORMAL"
    t_name = f"{sample_id}_tumor" if f"{sample_id}_tumor" in record.samples else "TUMOR"

    n_fmt, t_fmt = record.samples.get(n_name, {}), record.samples.get(t_name, {})
    n_dp, t_dp = get_safe_format_field(n_fmt, "DP", 0), get_safe_format_field(t_fmt, "DP", 0)
    n_ad, t_ad = n_fmt.get("AD", (0, 0)), t_fmt.get("AD", (0, 0))

    n_alt = n_ad[alt_allele_index + 1] if isinstance(n_ad, (list, tuple)) and len(n_ad) > alt_allele_index + 1 else 0
    t_alt = t_ad[alt_allele_index + 1] if isinstance(t_ad, (list, tuple)) and len(t_ad) > alt_allele_index + 1 else 0
    n_ref = n_ad[0] if isinstance(n_ad, (list, tuple)) else 0
    t_ref = t_ad[0] if isinstance(t_ad, (list, tuple)) else 0
    
    n_vaf = n_alt / n_dp if n_dp > 0 else 0
    t_vaf = t_alt / t_dp if t_dp > 0 else 0

    all_csq = [dict(zip(csq_header, e.split("|"))) for e in record.info.get("CSQ", [])]
    allele_csq = [e for e in all_csq if e.get("Allele") == alt_allele]
    
    if allele_csq:
        canonical = sorted(allele_csq, key=lambda csq: (get_most_severe_consequence_rank(csq), csq.get("CANONICAL") != "YES"))[0]
    else: 
        canonical = {}

    hgvsc = canonical.get("HGVSc", "/")
    hgvsp = canonical.get("HGVSp", "/")
    
    return {
        "VariantKey": variant_key, "CHROM": chrom, "POS": pos, "REF": ref, "ALT": alt_allele,
        "Normal_DP": n_dp, "Normal_AD": f"{n_ref},{n_alt}", "Normal_VAF": f"{n_vaf:.4f}",
        "Tumor_DP": t_dp, "Tumor_AD": f"{t_ref},{t_alt}", "Tumor_VAF": f"{t_vaf:.4f}",
        "TLOD": get_safe_info_field(record.info, "TLOD", 0.0), 
        "gnomAD_AF": get_safe_info_field(record.info, "gnomAD_AF", "/"),
        "COSMIC_CNT": get_safe_info_field(record.info, "COSMIC_CNT", 0), 
        "Gene": canonical.get("SYMBOL", "/"),
        "Transcript": canonical.get("Feature", "/"), 
        "HGVSc": hgvsc.split(':')[-1] if ':' in hgvsc else hgvsc,
        "HGVSp": hgvsp.split(':')[-1] if ':' in hgvsp else hgvsp, 
        "Consequence": canonical.get("Consequence", "/"),
        "Impact": canonical.get("IMPACT", "/"), 
        "Coding_ratio": calculate_coding_ratio(allele_csq), 
        "AA_chaged_ratio": calculate_AAchaged_ratio(allele_csq),
        "all_csq": allele_csq
    }

# [FIX] 增加类型强转安全保护函数
def _to_float(val, default=0.0):
    try:
        return float(val) if val not in ("/", None, "") else default
    except (ValueError, TypeError):
        return default

def classify_variant_new(variant, params):
    """根据黑名单、白名单、RNA捞回机制对变异进行分级"""
    p = params
    
    # [FIX] 使用稳健的浮点数转换
    n_vaf, t_vaf = _to_float(variant["Normal_VAF"]), _to_float(variant["Tumor_VAF"])
    n_dp, t_dp = variant.get("Normal_DP", 0), variant.get("Tumor_DP", 0)
    
    try:
        n_vd = int(variant["Normal_AD"].split(",")[1])
        t_vd = int(variant["Tumor_AD"].split(",")[1])
    except (IndexError, ValueError, AttributeError):
        n_vd, t_vd = 0, 0

    gnomad_af = _to_float(variant.get("gnomAD_AF", "/"), 0.0)
    tlod = _to_float(variant.get("TLOD", 0.0), 0.0)
    
    consequences = variant.get("Consequence", "").split('&')
    rna_support_reads = int(variant.get("RNA_Support_Reads", 0))
    gene = variant.get("Gene", "/")
    hgvsp = variant.get("HGVSp", "/")
    
    details = []
    sub_status = get_amino_acid_sub_status(consequences)

    # === 1. 优先判定 Germline / LOH ===
    if n_vaf >= p.get('min_normal_vaf_germline', 0.18) and n_dp >= p.get('min_normal_dp_germline', 10):
        details.append(f"Germline_Signal(N_VAF={n_vaf:.2f})")
        if (0.3 <= n_vaf <= 0.7) and (t_vaf > 0.9 or t_vaf < 0.1):
            return "LOH", sub_status, f"LOH_Shift(T_VAF={t_vaf:.2f})"
        return "Germline", sub_status, "; ".join(details)

    # === 2. 黑名单拦截 (Blacklist) ===
    if is_blacklisted(gene):
        return "Blacklist_Noise", sub_status, "Matched_Blacklist_Artifact"

    # === 3. 白名单保送 (Whitelist) ===
    if is_whitelisted(gene, hgvsp):
        details.append("[Whitelisted_Driver_Rescue]")
        return "Somatic", sub_status, "; ".join(details)

    # === 4. 常规噪音判定 (Noise) ===
    is_noise = False
    noise_reasons = []
    if t_vd < p.get('min_tumor_vd_somatic', 3):
        is_noise = True; noise_reasons.append(f"Low_T_VD({t_vd})")
    if t_vaf < p.get('min_tumor_vaf_somatic', 0.02):
        is_noise = True; noise_reasons.append(f"Low_T_VAF({t_vaf:.2f})")
    if n_vd > p.get('max_normal_vd_somatic', 3):
        is_noise = True; noise_reasons.append(f"High_N_VD_Cont({n_vd})")
    if tlod < p.get('min_tlod_somatic', 6.0):
        is_noise = True; noise_reasons.append(f"Low_TLOD({tlod})")
    if gnomad_af > p.get('max_gnomad_af_somatic', 0.001):
        is_noise = True; noise_reasons.append(f"High_gnomAD_AF({gnomad_af:.4f})")
        
    # === 5. RNA捞回机制 (Rescue) ===
    if is_noise:
        if rna_support_reads >= p.get('min_rna_reads_for_rescue', 2):
            details.append(f"Rescued_from_Noise(RNA_Reads={rna_support_reads})")
            if n_vaf >= p.get('min_rescue_n_vaf_germline', 0.1):
                return "Germline", sub_status, "Classified_as_Germline_Post_Rescue"
            return "Somatic", sub_status, "; ".join(details)
        return "Noise", sub_status, "; ".join(noise_reasons)

    # === 6. 核心通过 ===
    details.append("Pass_Somatic_Filters")
    return "Somatic", sub_status, "; ".join(details)

def load_auxiliary_data(config):
    logging.info("Loading auxiliary data...")
    varscan_variants, rna_support = set(), {}
    
    if config.input.get('varscan_vcf'):
        try:
            with pysam.VariantFile(config.input['varscan_vcf']) as vcf:
                for r in vcf: 
                    varscan_variants.add(f"{r.chrom}-{r.pos}-{r.ref}-{r.alts[0]}")
        except Exception as e: 
            logging.warning(f"Failed to load VarScan: {e}")
    
    if os.path.exists(config.input['rna_counts_tumor']):
        try:
            rna_df = pd.read_csv(config.input['rna_counts_tumor'], sep='\t')
            rna_support = {f"{r['contig']}-{r['position']}-{r['refAllele']}-{r['altAllele']}": r['altCount'] for _, r in rna_df.iterrows()}
        except Exception as e: 
            logging.warning(f"Failed to load RNA counts: {e}")
        
    return varscan_variants, rna_support

def parse_and_classify_variants(config, varscan_variants, rna_support):
    logging.info("Parsing and classifying variants from primary VCF...")
    all_variants = []
    with pysam.VariantFile(config.input['somatic_vcf']) as vcf:
        csq_header = get_csq_header(vcf)
        for record in vcf:
            for i in range(len(record.alts)):
                try: 
                    parsed = parse_vcf_record(record, i, csq_header, config.params['sample_id'])
                    all_variants.append(parsed)
                except Exception as e: 
                    logging.warning(f"Could not parse record {record.chrom}:{record.pos}: {e}")

    if not all_variants: 
        return pd.DataFrame()

    df = pd.DataFrame(all_variants)
    df["VarScan_Support"] = df["VariantKey"].isin(varscan_variants)
    df['RNA_Support_Reads'] = df['VariantKey'].map(rna_support).fillna(0).astype(int)

    classification_results = df.apply(lambda r: classify_variant_new(r, config.params['somatic_tiering_params']), axis=1)
    df[["Primary_Status", "Sub_Status", "Evidence_Details"]] = pd.DataFrame(classification_results.tolist(), index=df.index)
    
    # 新增标签计算 (Whitelist / Aware)
    df['Whitelist_Driver'] = df.apply(lambda r: "Yes" if is_whitelisted(r['Gene'], r['HGVSp']) else "/", axis=1)
    df['Biological_Aware'] = df['Gene'].map(AWARE_MAP).fillna("/")
    
    return df

def parse_hgvsp_for_aa_pos(hgvsp_str):
    if not isinstance(hgvsp_str, str) or not hgvsp_str.startswith('p.'): return None
    match = re.search(r'p\.[A-Z][a-z]{2}(\d+)', hgvsp_str)
    return int(match.group(1)) if match else None

def find_companion_germline_map(df):
    logging.info("Finding nearby germline variants using protein-aware method...")
    protein_map_list = []
    valid_rows = df[df['all_csq'].apply(lambda x: isinstance(x, list) and len(x) > 0)]
    
    for _, row in valid_rows.iterrows():
        if pd.isna(row.get('Primary_Status')): continue
        for csq in row['all_csq']:
            hgvsp = csq.get('HGVSp')
            if hgvsp and ':' in hgvsp:
                protein_id = hgvsp.split(':')[0]
                aa_pos = parse_hgvsp_for_aa_pos(hgvsp.split(':')[1])
                if protein_id and aa_pos is not None:
                    protein_map_list.append({
                        'VariantKey': row['VariantKey'], 'Primary_Status': row['Primary_Status'],
                        'Sub_Status': row['Sub_Status'], 'protein_id': protein_id, 'aa_pos': aa_pos
                    })

    if not protein_map_list: return {}
    pm_df = pd.DataFrame(protein_map_list)
    
    somatic_map = pm_df[(pm_df['Primary_Status'] == 'Somatic') & (pm_df['Sub_Status'] != 'No_AA_Altered')]
    germline_map = pm_df[(pm_df['Primary_Status'] == 'Germline') & (pm_df['Sub_Status'] != 'No_AA_Altered')]
    if somatic_map.empty or germline_map.empty: return {}

    merged = pd.merge(somatic_map, germline_map, on='protein_id', suffixes=('_somatic', '_germline'))
    companions = merged[abs(merged['aa_pos_somatic'] - merged['aa_pos_germline']) <= 18]
    
    companion_dict = defaultdict(set)
    for _, row in companions.iterrows():
        companion_dict[row['VariantKey_germline']].add(row['VariantKey_somatic'])
    return {k: list(v) for k, v in companion_dict.items()}

def generate_sample_warnings(df):
    """分析数据表，生成全局级别的临床/生物学预警信息"""
    warnings = []
    aware_muts = df[(df['Biological_Aware'] != "/") & 
                    (df['Primary_Status'].isin(['Somatic', 'LOH'])) & 
                    (df['Sub_Status'] != 'No_AA_Altered')]
    
    for _, row in aware_muts.iterrows():
        gene = row['Gene']
        cat = row['Biological_Aware']
        impact = row['Impact']
        
        if 'APM_Loss' in cat:
            warnings.append(f"CRITICAL: {gene} mutation ({impact}) detected! Potential HLA class I presentation defect.")
        elif 'Splice_Altered' in cat:
            warnings.append(f"INFO: {gene} splicing factor mutation detected. Consider exploring alternative splicing-derived neoantigens.")
        elif 'MMR_Defect' in cat or 'Polymerase_Defect' in cat:
            warnings.append(f"INFO: {gene} ({cat}) mutation detected. Patient may exhibit hypermutated phenotype.")
            
    if not warnings:
        warnings.append("No critical immune evasion or structural warnings detected.")
    return warnings

def generate_reports(df, companion_map, config):
    logging.info("Generating Excel reports...")
    output_filtered_xlsx = config.output['xlsx_report_filter']
    output_neopeptides_xlsx = config.output['xlsx_report_Somatic']

    # 定义颜色样式
    germline_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 浅黄
    whitelist_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # 浅绿

    df['Manual_Select'] = 'no'
    df['Companion_To'] = '/'
    
    # 处理 Somatic 的 Manual_Select
    somatic_yes_mask = (df['Primary_Status'] == 'Somatic') & ((df['RNA_Support_Reads'] > 0) | (df['Whitelist_Driver'] == 'Yes'))
    df.loc[somatic_yes_mask, 'Manual_Select'] = 'yes'
    somatic_manual_map = df[df['Primary_Status'] == 'Somatic'].set_index('VariantKey')['Manual_Select'].to_dict()

    # 处理 Germline 及其伴侣状态
    for g_key, s_partners in companion_map.items():
        g_idx = df[df['VariantKey'] == g_key].index
        if not g_idx.empty:
            df.loc[g_idx, 'Companion_To'] = ";".join(s_partners)
            if any(somatic_manual_map.get(s) == 'yes' for s in s_partners):
                df.loc[g_idx, 'Manual_Select'] = 'yes'

    # 定义导出的最终列
    final_cols = [
        "VariantKey", "CHROM", "POS", "REF", "ALT", "Gene", "Transcript", "HGVSc", "HGVSp", 
        "Consequence", "Impact", "Coding_ratio", "AA_chaged_ratio", "Manual_Select", 
        "Whitelist_Driver", "Biological_Aware", "VarScan_Support", "RNA_Support_Reads", 
        "Normal_DP", "Normal_AD", "Normal_VAF", "Tumor_DP", "Tumor_AD", "Tumor_VAF", 
        "TLOD", "gnomAD_AF", "COSMIC_CNT", "Primary_Status", "Sub_Status", 
        "Evidence_Details", "Companion_To"
    ]
    available_cols = [c for c in final_cols if c in df.columns]

    # 初始化分类存放容器
    somatic_sheets = defaultdict(list)
    filtered_sheets = {} 

    # --- 数据分流逻辑 ---
    filtered_sheets['Noise'] = df[df['Primary_Status'] == 'Noise']
    filtered_sheets['Blacklist_Filtered'] = df[df['Primary_Status'] == 'Blacklist_Noise']
    filtered_sheets['LOH'] = df[df['Primary_Status'] == 'LOH']
    filtered_sheets['Somatic_without_AA_altered'] = df[(df['Primary_Status'] == 'Somatic') & (df['Sub_Status'] == 'No_AA_Altered')]
    filtered_sheets['Germline_without_AA_altered'] = df[(df['Primary_Status'] == 'Germline') & (df['Sub_Status'] == 'No_AA_Altered')]

    somatic_aa = df[(df['Primary_Status'] == 'Somatic') & (df['Sub_Status'] != 'No_AA_Altered')]
    somatic_sheets['low scale variants'].append(somatic_aa[somatic_aa['Sub_Status'].isin(['Single_AA_Altered', 'Stop_Gained'])])
    somatic_sheets['Large scale variants'].append(somatic_aa[somatic_aa['Sub_Status'] == 'Long_AA_Altered'])

    germline_aa = df[(df['Primary_Status'] == 'Germline') & (df['Sub_Status'] != 'No_AA_Altered')]
    isolated_germlines = []
    
    for _, row in germline_aa.iterrows():
        if row['VariantKey'] in companion_map:
            p_stats = df[df['VariantKey'].isin(companion_map[row['VariantKey']])]['Sub_Status'].tolist()
            target = 'Large scale variants' if any(s == 'Long_AA_Altered' for s in p_stats) else 'low scale variants'
            somatic_sheets[target].append(pd.DataFrame([row]))
        else: 
            isolated_germlines.append(row)
    
    if isolated_germlines: 
        filtered_sheets['Germline_with_AA_altered'] = pd.DataFrame(isolated_germlines)

    # 排序辅助函数：Whitelist 优先，其次 RNA 支持降序，最后 VAF 降序
    def sort_somatic_df(df_to_sort):
        if df_to_sort.empty: return df_to_sort
        df_to_sort['sort_wl'] = df_to_sort['Whitelist_Driver'].apply(lambda x: 0 if x == 'Yes' else 1)
        df_to_sort['sort_vaf'] = pd.to_numeric(df_to_sort['Tumor_VAF'], errors='coerce').fillna(0)
        sorted_df = df_to_sort.sort_values(by=['sort_wl', 'RNA_Support_Reads', 'sort_vaf'], ascending=[True, False, False])
        return sorted_df.drop(columns=['sort_wl', 'sort_vaf']).reset_index(drop=True)

    # --- 1. 写入 Somatic 报告 (含 Summary Sheet) ---
    with pd.ExcelWriter(output_neopeptides_xlsx, engine='openpyxl') as writer:
        
        # A. 生成 Cover Page (Sample_Summary)
        wb = writer.book
        ws_summary = wb.create_sheet(title="Sample_Summary", index=0)
        
        warnings = generate_sample_warnings(df)
        ws_summary.append(["Sample Neoantigen Quality & Warning Report"])
        ws_summary.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws_summary.append([])
        
        ws_summary.append(["Total Somatic Mutations (AA altered):", len(somatic_aa)])
        ws_summary.append(["Total RNA Rescued Somatics:", len(somatic_aa[somatic_aa['RNA_Support_Reads'] > 0])])
        ws_summary.append(["Total Whitelisted Hotspots:", len(somatic_aa[somatic_aa['Whitelist_Driver'] == 'Yes'])])
        ws_summary.append([])
        
        ws_summary.append(["=== BIOLOGICAL WARNINGS ==="])
        for w in warnings:
            ws_summary.append([w])
            if w.startswith("CRITICAL"):
                ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(color="FF0000", bold=True)
                
        ws_summary.column_dimensions['A'].width = 100

        # B. 写入变异数据并填色
        for sheet_name, df_list in somatic_sheets.items():
            if df_list:
                merged_df = pd.concat(df_list).drop(columns=['all_csq'], errors='ignore')
                if not merged_df.empty:
                    final_df = sort_somatic_df(merged_df.reindex(columns=available_cols).fillna('/'))
                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    ws = writer.sheets[sheet_name]
                    for i, row in final_df.iterrows():
                        excel_row = i + 2
                        is_germline_companion = row['Primary_Status'] == 'Germline' and row.get('Companion_To', '/') != '/'
                        is_whitelisted_driver = row['Whitelist_Driver'] == 'Yes'
                        
                        for col_idx in range(1, len(available_cols) + 1):
                            cell = ws.cell(row=excel_row, column=col_idx)
                            if is_whitelisted_driver:
                                cell.fill = whitelist_fill # 白名单浅绿色
                            elif is_germline_companion:
                                cell.fill = germline_fill # 伴侣Germline浅黄色

    # --- 2. 写入 Filtered 报告 ---
    with pd.ExcelWriter(output_filtered_xlsx, engine='openpyxl') as writer:
        for sheet_name, df_data in filtered_sheets.items():
            if not df_data.empty:
                final_df = df_data.drop(columns=['all_csq'], errors='ignore').reindex(columns=available_cols).fillna('/').sort_values(by=["CHROM", "POS"])
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)

def main():
    config = get_args()
    log_params = {"level": logging.INFO, "format": "%(asctime)s - %(levelname)s - %(message)s", "force": True}
    if config.log: 
        log_params["filename"] = config.log[0]
    else: 
        log_params["stream"] = sys.stdout
    logging.basicConfig(**log_params)

    varscan_variants, rna_support = load_auxiliary_data(config)
    classified_df = parse_and_classify_variants(config, varscan_variants, rna_support)
    
    if classified_df.empty:
        logging.info("No variants to process. Exiting.")
        return

    companion_map = find_companion_germline_map(classified_df)
    generate_reports(classified_df, companion_map, config)
    logging.info("Script finished successfully.")

if __name__ == "__main__":
    main()