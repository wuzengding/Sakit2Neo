#!/usr/bin/env python3
# generate_rna_somatic_report.py
"""
Generate a comprehensive, annotated Excel report for RNA-based somatic variants.

Key Features:
- Classifies variants from RNA-Mutect2 VCF.
- Uses Tumor AD (Allele Depth) directly as RNA Support Reads.
- Handles multi-allelic sites.
- Identifies germline variants nearby.
- Produces polished Excel reports.
"""

import pandas as pd
import pysam
import logging
import argparse
import sys
import os
import re
import yaml
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

try:
    import pyensembl
except ImportError:
    pyensembl = None

# 复用原本的常量定义
CONSEQUENCE_SEVERITY_ORDER = [
    'transcript_ablation', 'splice_acceptor_variant', 'splice_donor_variant',
    'stop_gained', 'frameshift_variant', 'stop_lost', 'start_lost',
    'transcript_amplification', 'inframe_insertion', 'inframe_deletion',
    'missense_variant', 'protein_altering_variant', 'splice_donor_5th_base_variant',
    'splice_region_variant', 'splice_donor_region_variant',
    'splice_polypyrimidine_tract_variant', 'incomplete_terminal_codon_variant',
    'start_retained_variant', 'stop_retained_variant', 'synonymous_variant',
    'coding_sequence_variant', 'mature_miRNA_variant', '5_prime_UTR_variant',
    '3_prime_UTR_variant', 'non_coding_transcript_exon_variant', 'intron_variant',
    'NMD_transcript_variant', 'non_coding_transcript_variant',
    'coding_transcript_variant', 'upstream_gene_variant', 'downstream_gene_variant',
    'TFBS_ablation', 'TFBS_amplification', 'TF_binding_site_variant',
    'regulatory_region_ablation', 'regulatory_region_amplification',
    'regulatory_region_variant', 'feature_elongation', 'feature_truncation',
    'intergenic_variant', 'sequence_variant'
]

CONSEQUENCE_SEVERITY_MAP = {
    consequence: i for i, consequence in enumerate(CONSEQUENCE_SEVERITY_ORDER)
}

PROTEIN_ALTERING_CONSEQUENCES = {
    'transcript_ablation', 'splice_acceptor_variant', 'splice_donor_variant',
    'stop_gained', 'frameshift_variant', 'stop_lost', 'start_lost',
    'inframe_insertion', 'inframe_deletion', 'missense_variant',
    'protein_altering_variant',
}

def get_args():
    if 'snakemake' in globals():
        logging.info("Running in Snakemake mode.")
        if not pyensembl: raise ImportError("pyensembl is required but not installed.")
        return snakemake

    logging.info("Running in Standalone mode.")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--somatic-vcf", required=True, help="The RNA Mutect2 VEP-annotated VCF.")
    parser.add_argument("--varscan-vcf", required=False, help="[OPTIONAL] VarScan2 VCF.")
    # Removed --rna-counts-tumor as it's inherent in the VCF
    parser.add_argument("--cancer-genes", required=True)
    parser.add_argument("--gtf", required=True)
    parser.add_argument("--fasta", required=True)
    parser.add_argument("--output-xlsx-filter", required=True)
    parser.add_argument("--output-xlsx-somatic", required=True)
    parser.add_argument("--sample-id", required=True)
    parser.add_argument("--tiering-params-yaml", required=True)
    parser.add_argument("--log", help="Log file path.")
    args = parser.parse_args()

    class MockSnakemake:
        def __init__(self):
            self.input = {
                'somatic_vcf': args.somatic_vcf,
                'varscan_vcf': args.varscan_vcf if args.varscan_vcf else [],
                'cancer_genes': args.cancer_genes,
                'gtf': args.gtf, 'fasta': args.fasta
            }
            self.output = {
                'xlsx_report_filter_rna': args.output_xlsx_filter,
                'xlsx_report_Somatic_rna': args.output_xlsx_somatic
            }
            self.params = {'sample_id': args.sample_id}
            self.log = [args.log] if args.log else None
            with open(args.tiering_params_yaml, 'r') as f:
                config_yaml = yaml.safe_load(f)
                self.params['somatic_tiering_params'] = config_yaml.get("somatic_tiering")
    return MockSnakemake()

# --- Helper Functions (reused) ---
def get_csq_header(vcf_reader):
    if "CSQ" not in vcf_reader.header.info:
        raise ValueError("CSQ field not found in VCF header.")
    match = re.search(r"Format: ([\w|]+)", vcf_reader.header.info["CSQ"].description)
    if match: return match.group(1).split("|")
    raise ValueError("Could not parse CSQ header format string.")

def get_safe_info_field(info, key, default):
    if key in info:
        val = info.get(key)
        return val[0] if isinstance(val, (list, tuple)) else val
    return default

def get_safe_format_field(sample_format, key, default):
    val = sample_format.get(key)
    if val is None: return default
    return val[0] if isinstance(val, (list, tuple)) else val

def calculate_coding_ratio(csq_entries):
    if not csq_entries: return "0//0"
    total_transcripts = len(csq_entries)
    coding_transcripts = sum(1 for e in csq_entries if e.get("BIOTYPE") == "protein_coding")
    return f"{coding_transcripts}//{total_transcripts}"

def calculate_AAchaged_ratio(csq_entries):
    if not csq_entries: return "0//0"
    total_transcripts = len(csq_entries)
    protein_altering_count = 0
    for e in csq_entries:
        consequences = e.get("Consequence", "").split('&')
        if any(c in PROTEIN_ALTERING_CONSEQUENCES for c in consequences):
            protein_altering_count += 1
    return f"{protein_altering_count}//{total_transcripts}"

def get_most_severe_consequence_rank(csq_entry):
    consequences = csq_entry.get("Consequence", "").split('&')
    min_rank = len(CONSEQUENCE_SEVERITY_ORDER) 
    for c in consequences:
        rank = CONSEQUENCE_SEVERITY_MAP.get(c, min_rank)
        if rank < min_rank: min_rank = rank
    return min_rank

def get_amino_acid_sub_status(consequences):
    is_protein_altering = any(c in PROTEIN_ALTERING_CONSEQUENCES for c in consequences)
    if not is_protein_altering: return "No_AA_Altered"
    if 'stop_gained' in consequences: return "Stop_Gained"
    elif any(c in ['frameshift_variant', 'splice_acceptor_variant', 'splice_donor_variant', 'stop_lost'] for c in consequences):
        return "Long_AA_Altered"
    else: return "Single_AA_Altered"

# --- Parsing Logic (Modified for RNA) ---
def parse_vcf_record(record, alt_allele_index, csq_header, sample_id):
    alt_allele = record.alts[alt_allele_index]
    chrom, pos, ref = record.chrom, record.pos, record.ref
    variant_key = f"{chrom}-{pos}-{ref}-{alt_allele}"
    info = record.info

    # DNA VCF通常标记为 TUMOR/NORMAL 或 {sample}_tumor/{sample}_normal
    # RNA VCF (Mutect2) 也会遵循类似的命名，或者直接是 SAMPLE
    if f"{sample_id}_normal" in record.samples:
        n_name, t_name = f"{sample_id}_normal", f"{sample_id}_tumor"
    elif "NORMAL" in record.samples and "TUMOR" in record.samples:
        n_name, t_name = "NORMAL", "TUMOR"
    else:
        # 如果是单样本RNA分析 (tumor-only)，可能只有一个样本
        # 这里为了兼容代码结构，我们假设 record.samples[0] 是 Tumor
        t_name = record.samples[0].name
        n_name = None # No normal

    # 提取 Tumor 数据
    t_fmt = record.samples.get(t_name, {})
    t_dp = get_safe_format_field(t_fmt, "DP", 0)
    t_ad = t_fmt.get("AD", (0, 0))
    
    if isinstance(t_ad, (list, tuple)) and len(t_ad) > alt_allele_index + 1:
        t_alt_reads = t_ad[alt_allele_index + 1]
        t_ref_reads = t_ad[0]
    else: 
        t_alt_reads, t_ref_reads = 0, 0
    t_vaf = t_alt_reads / t_dp if t_dp > 0 else 0

    # 提取 Normal 数据 (如果存在)
    if n_name:
        n_fmt = record.samples.get(n_name, {})
        n_dp = get_safe_format_field(n_fmt, "DP", 0)
        n_ad = n_fmt.get("AD", (0, 0))
        if isinstance(n_ad, (list, tuple)) and len(n_ad) > alt_allele_index + 1:
            n_alt_reads = n_ad[alt_allele_index + 1]
            n_ref_reads = n_ad[0]
        else: n_alt_reads, n_ref_reads = 0, 0
        n_vaf = n_alt_reads / n_dp if n_dp > 0 else 0
    else:
        # Tumor-only mode or missing normal
        n_dp, n_alt_reads, n_ref_reads, n_vaf = 0, 0, 0, 0.0

    # CSQ Parsing
    all_csq = [dict(zip(csq_header, e.split("|"))) for e in info.get("CSQ", [])]
    allele_csq = [e for e in all_csq if e.get("Allele") == alt_allele]
    
    if not allele_csq: canonical = {}
    else:
        sorted_csq = sorted(
            allele_csq,
            key=lambda csq: (
                get_most_severe_consequence_rank(csq),
                csq.get("CANONICAL") != "YES"
            )
        )
        canonical = sorted_csq[0]

    hgvsc = canonical.get("HGVSc", "/")
    hgvsp = canonical.get("HGVSp", "/")
    
    return {
        "VariantKey": variant_key, "CHROM": chrom, "POS": pos, "REF": ref, "ALT": alt_allele,
        "Normal_DP": n_dp, "Normal_AD": f"{n_ref_reads},{n_alt_reads}", "Normal_VAF": f"{n_vaf:.4f}",
        "Tumor_DP": t_dp, "Tumor_AD": f"{t_ref_reads},{t_alt_reads}", "Tumor_VAF": f"{t_vaf:.4f}",
        
        # === 关键修改 ===
        # 在 RNA 报告中，RNA_Support_Reads 直接等于 Tumor Alt Depth
        # 因为这个 VCF 本身就是来源于 RNA 测序
        "RNA_Support_Reads": t_alt_reads, 
        
        "TLOD": get_safe_info_field(info, "TLOD", 0.0), "gnomAD_AF": get_safe_info_field(info, "gnomAD_AF", 0.0),
        "COSMIC_CNT": get_safe_info_field(info, "COSMIC_CNT", 0), "Gene": canonical.get("SYMBOL", "/"),
        "Transcript": canonical.get("Feature", "/"), "HGVSc": hgvsc.split(':')[-1] if ':' in hgvsc else hgvsc,
        "HGVSp": hgvsp.split(':')[-1] if ':' in hgvsp else hgvsp, "Consequence": canonical.get("Consequence", "/"),
        "Impact": canonical.get("IMPACT", "/"), 
        "Coding_ratio": calculate_coding_ratio(allele_csq), "AA_chaged_ratio": calculate_AAchaged_ratio(allele_csq),
        "all_csq": allele_csq
    }

# --- Classification Logic (Reused but applied to RNA context) ---
def classify_variant_rna(variant, params):
    p = params
    n_vaf, t_vaf = float(variant["Normal_VAF"]), float(variant["Tumor_VAF"])
    n_dp = variant["Normal_DP"]
    t_vd = variant["RNA_Support_Reads"] # Reuse the field we populated
    n_vd = int(variant["Normal_AD"].split(",")[1])
    gnomad_af = float(variant["gnomAD_AF"]) if variant["gnomAD_AF"] != "/" else 0.0
    tlod = float(variant["TLOD"])
    consequences = variant.get("Consequence", "").split('&')
    
    details = []

    # 1. Germline Signal
    # 如果 Normal 样本也是 RNA，高 VAF 可能代表 Germline 或 RNA Editing
    # 如果 Normal 样本是 DNA，高 VAF 就是 Germline
    if n_vaf >= p.get('min_normal_vaf_germline', 0.18) and n_dp >= p.get('min_normal_dp_germline', 10):
        details.append(f"Germline_Signal(N_VAF={n_vaf:.2f})")
        if (0.3 <= n_vaf <= 0.7) and (t_vaf > 0.9 or t_vaf < 0.1):
            primary_status = "LOH"
            details.append(f"LOH_Shift(T_VAF={t_vaf:.2f})")
        else:
            primary_status = "Germline"
        sub_status = get_amino_acid_sub_status(consequences)
        return primary_status, sub_status, "; ".join(details)

    # 2. Noise Detection
    is_noise = False
    noise_reasons = []
    
    if t_vd < p.get('min_tumor_vd_somatic', 3):
        is_noise = True; noise_reasons.append(f"Low_T_VD({t_vd})")
    # RNA-seq VAF 波动很大，有些杂合突变因为表达量倾斜VAF可能很低，或者非常高
    if t_vaf < p.get('min_tumor_vaf_somatic', 0.02):
        is_noise = True; noise_reasons.append(f"Low_T_VAF({t_vaf:.2f})")
    if n_vd > p.get('max_normal_vd_somatic', 3):
        is_noise = True; noise_reasons.append(f"High_N_VD({n_vd})")
    if tlod < p.get('min_tlod_somatic', 6.0):
        is_noise = True; noise_reasons.append(f"Low_TLOD({tlod})")
    if gnomad_af > p.get('max_gnomad_af_somatic', 0.001):
        is_noise = True; noise_reasons.append(f"High_gnomAD_AF({gnomad_af:.4f})")

    # 3. Rescue Logic?
    # 在 DNA 流程中，我们用 RNA Support 来捞回。
    # 在 RNA 流程中，t_vd 本身就是 RNA reads。如果 t_vd 很低，就是真的低，没法捞。
    # 所以如果是 Noise，基本就是 Noise。
    # 唯一例外：如果 gnomAD 高，但 COSMIC 也有? (这里暂不复杂化，保持一致)
    
    if is_noise:
        return "Noise", None, "; ".join(noise_reasons)

    # 4. Somatic
    primary_status = "Somatic"
    details.append("Pass_Somatic_Filters")
    sub_status = get_amino_acid_sub_status(consequences)
    return primary_status, sub_status, "; ".join(details)

# --- Auxiliary Data (Simplified) ---
def load_auxiliary_data(config):
    logging.info("Loading auxiliary data (VarScan only)...")
    varscan_vcf_path = config.input.get('varscan_vcf')
    varscan_variants = set()
    if varscan_vcf_path and isinstance(varscan_vcf_path, str) and os.path.exists(varscan_vcf_path):
        with pysam.VariantFile(varscan_vcf_path) as vcf:
            for r in vcf:
                varscan_variants.add(f"{r.chrom}-{r.pos}-{r.ref}-{r.alts[0]}")
    return varscan_variants

# --- Processing Functions ---
def parse_and_classify_variants(config, varscan_variants):
    logging.info("Parsing and classifying variants from RNA VCF...")
    somatic_vcf_path = config.input['somatic_vcf']
    sample_id = config.params['sample_id']
    tiering_params = config.params['somatic_tiering_params']

    all_variants = []
    with pysam.VariantFile(somatic_vcf_path) as vcf:
        csq_header = get_csq_header(vcf)
        for record in vcf:
            for i in range(len(record.alts)):
                try:
                    parsed = parse_vcf_record(record, i, csq_header, sample_id)
                    all_variants.append(parsed)
                except Exception as e:
                    logging.warning(f"Could not parse record {record.chrom}:{record.pos}: {e}")
    
    if not all_variants: return pd.DataFrame()

    df = pd.DataFrame(all_variants)
    df["VarScan_Support"] = df["VariantKey"].isin(varscan_variants)
    
    classification_results = df.apply(
        lambda r: classify_variant_rna(r, tiering_params), axis=1
    )
    df[["Primary_Status", "Sub_Status", "Evidence_Details"]] = pd.DataFrame(
        classification_results.tolist(), index=df.index
    )
    return df

# --- Companion Logic (Same as DNA) ---
def parse_hgvsp_for_aa_pos(hgvsp_str):
    if not isinstance(hgvsp_str, str) or not hgvsp_str.startswith('p.'): return None
    match = re.search(r'p\.[A-Z][a-z]{2}(\d+)', hgvsp_str)
    return int(match.group(1)) if match else None

def find_companion_germline_map(df):
    protein_map_list = []
    valid_rows = df[df['all_csq'].apply(lambda x: isinstance(x, list) and len(x) > 0)]
    for _, row in valid_rows.iterrows():
        if pd.isna(row.get('Primary_Status')): continue
        for csq in row['all_csq']:
            hgvsp = csq.get('HGVSp')
            if hgvsp and ':' in hgvsp:
                protein_id = hgvsp.split(':')[0]
                aa_pos = parse_hgvsp_for_aa_pos(hgvsp.split(':')[1])
                if protein_id and aa_pos is not None:
                    protein_map_list.append({
                        'VariantKey': row['VariantKey'],
                        'Primary_Status': row['Primary_Status'],
                        'Sub_Status': row['Sub_Status'],
                        'protein_id': protein_id, 'aa_pos': aa_pos
                    })

    if not protein_map_list: return {}
    protein_map_df = pd.DataFrame(protein_map_list)
    
    somatic_map = protein_map_df[
        (protein_map_df['Primary_Status'] == 'Somatic') & 
        (protein_map_df['Sub_Status'] != 'No_AA_Altered')
    ]
    germline_map = protein_map_df[
        (protein_map_df['Primary_Status'] == 'Germline') & 
        (protein_map_df['Sub_Status'] != 'No_AA_Altered')
    ]

    if somatic_map.empty or germline_map.empty: return {}
    
    merged = pd.merge(somatic_map, germline_map, on='protein_id', suffixes=('_somatic', '_germline'))
    merged['distance'] = abs(merged['aa_pos_somatic'] - merged['aa_pos_germline'])
    companions = merged[merged['distance'] <= 18]
    
    if companions.empty: return {}

    companion_dict = defaultdict(set)
    for _, row in companions.iterrows():
        companion_dict[row['VariantKey_germline']].add(row['VariantKey_somatic'])
    return {k: list(v) for k, v in companion_dict.items()}

# --- Report Generation (Modified filenames) ---
def generate_reports(df, companion_germline_map, config):
    output_filtered_xlsx = config.output['xlsx_report_filter_rna']
    output_neopeptides_xlsx = config.output['xlsx_report_Somatic_rna']
    germline_highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    df['Manual_Select'] = 'no'
    df['Companion_To'] = '/'
    
    # RNA Manual Select logic: Since this IS RNA data, if it's Somatic, it has RNA support by definition
    # provided it passed the t_vd filters.
    somatic_yes_mask = (df['Primary_Status'] == 'Somatic')
    df.loc[somatic_yes_mask, 'Manual_Select'] = 'yes'
    somatic_manual_status_map = df[df['Primary_Status'] == 'Somatic'].set_index('VariantKey')['Manual_Select'].to_dict()

    for germline_key, somatic_partners in companion_germline_map.items():
        germline_idx = df[df['VariantKey'] == germline_key].index
        if not germline_idx.empty:
            df.loc[germline_idx, 'Companion_To'] = ";".join(somatic_partners)
            is_companion_yes = any(somatic_manual_status_map.get(k) == 'yes' for k in somatic_partners)
            if is_companion_yes: df.loc[germline_idx, 'Manual_Select'] = 'yes'

    final_cols = [
        "VariantKey", "CHROM", "POS", "REF", "ALT", 
        "Gene", "Transcript", "HGVSc", "HGVSp", 
        "Consequence", "Impact", 
        "Coding_ratio", "AA_chaged_ratio", "Manual_Select",
        "VarScan_Support", "RNA_Support_Reads", 
        "Normal_DP", "Normal_AD", "Normal_VAF", 
        "Tumor_DP", "Tumor_AD", "Tumor_VAF", 
        "TLOD", "gnomAD_AF", "COSMIC_CNT", 
        "Primary_Status", "Sub_Status", "Evidence_Details", "Companion_To"
    ]
    available_cols = [c for c in final_cols if c in df.columns]

    somatic_neopeptides_sheets = defaultdict(list)
    filtered_variants_sheets = defaultdict(list)

    filtered_variants_sheets['Noise'] = df[df['Primary_Status'] == 'Noise']
    filtered_variants_sheets['LOH'] = df[df['Primary_Status'] == 'LOH']
    filtered_variants_sheets['Somatic_without_AA_altered'] = df[(df['Primary_Status'] == 'Somatic') & (df['Sub_Status'] == 'No_AA_Altered')]
    filtered_variants_sheets['Germline_without_AA_altered'] = df[(df['Primary_Status'] == 'Germline') & (df['Sub_Status'] == 'No_AA_Altered')]

    somatic_aa_altered = df[(df['Primary_Status'] == 'Somatic') & (df['Sub_Status'] != 'No_AA_Altered')]
    somatic_neopeptides_sheets['low scale variants'].append(somatic_aa_altered[somatic_aa_altered['Sub_Status'].isin(['Single_AA_Altered', 'Stop_Gained'])])
    somatic_neopeptides_sheets['Large scale variants'].append(somatic_aa_altered[somatic_aa_altered['Sub_Status'] == 'Long_AA_Altered'])

    germline_aa_altered = df[(df['Primary_Status'] == 'Germline') & (df['Sub_Status'] != 'No_AA_Altered')]
    isolated_germlines = []
    for _, row in germline_aa_altered.iterrows():
        if row['VariantKey'] in companion_germline_map:
            somatic_partners = companion_germline_map[row['VariantKey']]
            partner_statuses = df[df['VariantKey'].isin(somatic_partners)]['Sub_Status'].tolist()
            target_sheet = 'Large scale variants' if any(s == 'Long_AA_Altered' for s in partner_statuses) else 'low scale variants'
            somatic_neopeptides_sheets[target_sheet].append(pd.DataFrame([row]))
        else: isolated_germlines.append(row)
    if isolated_germlines: filtered_variants_sheets['Germline_with_AA_altered'] = pd.DataFrame(isolated_germlines)

    with pd.ExcelWriter(output_neopeptides_xlsx, engine='openpyxl') as writer:
        for sheet_name, df_list in somatic_neopeptides_sheets.items():
            if df_list:
                final_df = pd.concat(df_list).drop(columns=['all_csq'], errors='ignore')
                final_df = final_df.reindex(columns=available_cols).fillna('/').sort_values(by=["CHROM", "POS"]).reset_index(drop=True)
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                for i, row in final_df.iterrows():
                    if row['Primary_Status'] == 'Germline' and row.get('Companion_To', '/') != '/':
                        for col_idx in range(1, len(available_cols) + 1):
                            ws.cell(row=i + 2, column=col_idx).fill = germline_highlight_fill

    with pd.ExcelWriter(output_filtered_xlsx, engine='openpyxl') as writer:
        for sheet_name, df_data in filtered_variants_sheets.items():
            if not df_data.empty:
                final_df = df_data.drop(columns=['all_csq'], errors='ignore')
                final_df = final_df.reindex(columns=available_cols).fillna('/').sort_values(by=["CHROM", "POS"])
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)

def main():
    config = get_args()
    log_file = config.log[0] if config.log else None
    log_params = {"level": logging.INFO, "format": "%(asctime)s - %(levelname)s - %(message)s", "force": True}
    if log_file: log_params["filename"] = log_file
    else: log_params["stream"] = sys.stdout
    logging.basicConfig(**log_params)

    varscan_variants = load_auxiliary_data(config)
    classified_df = parse_and_classify_variants(config, varscan_variants)
    
    if classified_df.empty:
        logging.info("No variants found.")
        # Create empty dummy files to satisfy Snakemake
        with pd.ExcelWriter(config.output['xlsx_report_filter_rna'], engine='openpyxl') as w: pass
        with pd.ExcelWriter(config.output['xlsx_report_Somatic_rna'], engine='openpyxl') as w: pass
        return

    companion_map = find_companion_germline_map(classified_df)
    generate_reports(classified_df, companion_map, config)

if __name__ == "__main__":
    main()