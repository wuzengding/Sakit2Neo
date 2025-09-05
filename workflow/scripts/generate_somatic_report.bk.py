#!/usr/bin/env python3

"""
Generate a comprehensive, annotated Excel report for somatic variants.

This script integrates data from multiple sources to create a multi-sheet
Excel file for reviewing somatic and germline variants. It gracefully handles
the optional inclusion of a VarScan2 VCF for cross-validation and is robust
to different sample naming conventions and field availability in VCF files.
"""

import pandas as pd
import pysam
import logging
import argparse
import sys
import os
import re
import yaml
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import pyensembl
except ImportError:
    pyensembl = None


def get_args():
    """
    Determines if running under Snakemake or standalone and returns
    an appropriate configuration object.
    """
    if 'snakemake' in globals():
        logging.info("Running in Snakemake mode.")
        if not pyensembl: raise ImportError("pyensembl is required but not installed.")
        return snakemake

    logging.info("Running in Standalone mode.")
    if not pyensembl: sys.exit("ERROR: pyensembl is required. Please `pip install pyensembl`.")
        
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("--somatic-vcf", required=True, help="Somatic VEP-annotated VCF.")
    parser.add_argument("--germline-vcf", required=True, help="Germline VEP-annotated VCF.")
    parser.add_argument("--varscan-vcf", required=False, help="[OPTIONAL] VarScan2 VCF for validation.")
    parser.add_argument("--rna-counts-tumor", required=True, help="Tumor RNA ASE counts TSV.")
    parser.add_argument("--cancer-genes", required=True, help="List of known cancer genes.")
    parser.add_argument("--gtf", required=True, help="Reference GTF for pyensembl.")
    parser.add_argument("--fasta", required=True, help="Reference FASTA (needed by rule, not directly by this script's pyensembl call).")
    parser.add_argument("--output-xlsx", required=True, help="Output Excel report path.")
    parser.add_argument("--sample-id", required=True, help="Sample identifier (e.g., 'PATIENT1').")
    parser.add_argument("--tiering-params-yaml", required=True, help="YAML with somatic tiering parameters.")
    parser.add_argument("--log", help="Log file path. Logs to console if not provided.")
    args = parser.parse_args()

    class MockSnakemake:
        def __init__(self):
            self.input = {
                'somatic_vcf': args.somatic_vcf, 'germline_vcf': args.germline_vcf,
                'varscan_vcf': args.varscan_vcf if args.varscan_vcf else [],
                'rna_counts_tumor': args.rna_counts_tumor, 'cancer_genes': args.cancer-genes,
                'gtf': args.gtf, 'fasta': args.fasta
            }
            self.output = {'xlsx_report': args.output_xlsx}
            self.params = {'sample_id': args.sample_id}
            self.log = [args.log] if args.log else None
            with open(args.tiering_params_yaml, 'r') as f:
                config_yaml = yaml.safe_load(f)
                self.params['somatic_tiering_params'] = config_yaml.get("somatic_tiering")
                if not self.params['somatic_tiering_params']:
                    sys.exit(f"ERROR: 'somatic_tiering' not found in {args.tiering_params_yaml}")
    return MockSnakemake()


def get_csq_header(vcf_reader):
    """Parses VEP CSQ header to get the field order."""
    if "CSQ" not in vcf_reader.header.info:
        raise ValueError("CSQ field not found in VCF header.")
    csq_header_str = vcf_reader.header.info["CSQ"].description
    match = re.search(r"Format: ([\w|]+)", csq_header_str)
    if match: return match.group(1).split("|")
    raise ValueError("Could not parse CSQ header format string.")


def get_safe_info_field(info, key, default):
    """
    Safely retrieves a single value from a VCF INFO field, checking for presence
    first to avoid pysam's 'Invalid header' error.
    """
    if key in info:
        val = info.get(key)
        if isinstance(val, (list, tuple)):
            return val[0]
        return val
    return default


def get_safe_format_field(sample_format, key, default):
    """Safely retrieves a single value from a VCF FORMAT field."""
    val = sample_format.get(key)
    if val is None: return default
    if isinstance(val, (list, tuple)): return val[0]
    return val


def parse_vcf_record(record, vcf_reader, csq_header, sample_id):
    """Parses a single VCF record, handling different sample naming conventions."""
    chrom, pos, ref, alt = record.chrom, record.pos, record.ref, record.alts[0]
    variant_key = f"{chrom}-{pos}-{ref}-{alt}"
    info = record.info

    if f"{sample_id}_normal" in record.samples:
        normal_sample_name, tumor_sample_name = f"{sample_id}_normal", f"{sample_id}_tumor"
    else:
        normal_sample_name, tumor_sample_name = "NORMAL", "TUMOR"

    normal_fmt = record.samples.get(normal_sample_name, {})
    tumor_fmt = record.samples.get(tumor_sample_name, {})

    n_dp = get_safe_format_field(normal_fmt, "DP", 0)
    t_dp = get_safe_format_field(tumor_fmt, "DP", 0)
    n_ad = normal_fmt.get("AD", (0, 0))
    t_ad = tumor_fmt.get("AD", (0, 0))

    n_ad = n_ad if isinstance(n_ad, (list, tuple)) and len(n_ad) >= 2 else (n_dp - sum(n_ad if isinstance(n_ad, tuple) else [n_ad]), sum(n_ad if isinstance(n_ad, tuple) else [n_ad])) if n_dp > 0 and n_ad is not None else (0,0)
    t_ad = t_ad if isinstance(t_ad, (list, tuple)) and len(t_ad) >= 2 else (t_dp - sum(t_ad if isinstance(t_ad, tuple) else [t_ad]), sum(t_ad if isinstance(t_ad, tuple) else [t_ad])) if t_dp > 0 and t_ad is not None else (0,0)
    
    n_ref, n_alt = n_ad[0], n_ad[1]
    t_ref, t_alt = t_ad[0], t_ad[1]

    n_vaf = n_alt / n_dp if n_dp > 0 else 0
    t_vaf = t_alt / t_dp if t_dp > 0 else 0
    
    csq_entries = [dict(zip(csq_header, entry.split("|"))) for entry in info.get("CSQ", [])]
    
    canonical_entry = next((e for e in csq_entries if e.get("CANONICAL") == "YES" and e.get("Feature", "").startswith("ENST")), None)
    if not canonical_entry and csq_entries:
        canonical_entry = next((e for e in csq_entries if e.get("BIOTYPE") == "protein_coding"), csq_entries[0])
    if not canonical_entry: canonical_entry = {}

    hgvsc = canonical_entry.get("HGVSc", "/")
    hgvsp = canonical_entry.get("HGVSp", "/")
    
    return {
        "VariantKey": variant_key, "CHROM": chrom, "POS": pos, "ID": record.id, "REF": ref, "ALT": alt,
        "Normal_DP": n_dp, "Normal_AD": f"{n_ref},{n_alt}", "Normal_VAF": f"{n_vaf:.4f}",
        "Tumor_DP": t_dp, "Tumor_AD": f"{t_ref},{t_alt}", "Tumor_VAF": f"{t_vaf:.4f}",
        "TLOD": get_safe_info_field(info, "TLOD", 0.0),
        "gnomAD_AF": get_safe_info_field(info, "gnomAD_AF", 0.0),
        "COSMIC_CNT": get_safe_info_field(info, "COSMIC_CNT", 0),
        "Gene": canonical_entry.get("SYMBOL", "/"),
        "Transcript": canonical_entry.get("Feature", "/"),
        "HGVSc": hgvsc.split(':')[-1] if ':' in hgvsc else hgvsc,
        "HGVSp": hgvsp.split(':')[-1] if ':' in hgvsp else hgvsp,
        "Consequence": canonical_entry.get("Consequence", "/"),
        "Impact": canonical_entry.get("IMPACT", "/"),
    }


def tier_somatic_variant(variant_data, rna_support, varscan_support, params):
    """Applies a multi-tiered filtering logic to classify somatic variants."""
    t_vaf, n_vaf = float(variant_data["Tumor_VAF"]), float(variant_data["Normal_VAF"])
    t_dp, t_alt_reads = variant_data["Tumor_DP"], int(variant_data["Tumor_AD"].split(',')[1])
    tlod, rna_alt_reads = variant_data["TLOD"], rna_support.get('alt_reads', 0)

    if n_vaf >= params['germline_flags']['min_normal_vaf_for_germline']:
        return "Likely Germline", "High Normal VAF"
    
    p1 = params['tier1']
    if all([t_vaf >= p1['min_tumor_vaf'], n_vaf <= p1['max_normal_vaf'],
            t_dp >= p1['min_tumor_depth'], t_alt_reads >= p1['min_tumor_alt_reads'],
            tlod >= p1['min_tlod'], rna_alt_reads >= p1['min_rna_alt_reads'],
            (not p1['require_varscan_validation'] or varscan_support is True)]):
        return "Tier 1: High Confidence", "Passes all Tier 1 criteria"

    p2 = params['tier2']
    if all([t_vaf >= p2['min_tumor_vaf'], n_vaf <= p2['max_normal_vaf'],
            t_dp >= p2['min_tumor_depth'], t_alt_reads >= p2['min_tumor_alt_reads'],
            tlod >= p2['min_tlod']]):
        return "Tier 2: Medium Confidence", "Passes Tier 2 criteria, fails Tier 1"
    
    return "Tier 3: Low Confidence", "Fails Tier 1 & 2 criteria"


def find_nearby_germline_exon_aware(somatic_row, germline_by_chrom, ensembl_data):
    """Finds germline variants within 54 exonic bp of a somatic variant."""
    chrom, pos, tx_id = somatic_row["CHROM"], somatic_row["POS"], somatic_row["Transcript"]
    if not tx_id or not tx_id.startswith("ENST"): return []
    try:
        transcript = ensembl_data.transcript_by_id(tx_id.split('.')[0])
        somatic_offset = transcript.spliced_offset(pos)
    except Exception:
        logging.warning(f"Could not find transcript {tx_id} for variant at {chrom}:{pos}")
        return []
    
    nearby_germline = []
    for germline_record in germline_by_chrom.get(chrom, []):
        if abs(germline_record.pos - pos) > 10000: continue
        try:
            if transcript.contains(chrom, germline_record.pos):
                germline_offset = transcript.spliced_offset(germline_record.pos)
                if abs(somatic_offset - germline_offset) <= 54:
                    nearby_germline.append(germline_record)
        except Exception: continue
    return nearby_germline


def main():
    config = get_args()
    log_file = config.log[0] if config.log else None
    log_params = {"level": logging.INFO, "format": "%(asctime)s - %(levelname)s - %(message)s", "force": True}
    if log_file: log_params["filename"] = log_file
    else: log_params["stream"] = sys.stdout
    logging.basicConfig(**log_params)

    somatic_vcf, germline_vcf = config.input['somatic_vcf'], config.input['germline_vcf']
    varscan_vcf = config.input['varscan_vcf']
    rna_counts, cancer_genes_file = config.input['rna_counts_tumor'], config.input['cancer_genes']
    gtf, fasta = config.input['gtf'], config.input['fasta']
    output_xlsx, sample_id, tiering_params = config.output['xlsx_report'], config.params['sample_id'], config.params['somatic_tiering_params']
    
    logging.info("Initializing pyensembl...")
    
    # --- CORRECTED: Robustly locate the project's resources directory ---
    # Get the directory where this script is located
    # __file__ is a special variable that holds the path to the current script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Infer the project root, which is two levels up from this script (workflow/scripts/)
    project_root = os.path.dirname(os.path.dirname(script_dir))
    # Define the pyensembl cache path relative to the inferred project root
    pyensembl_cache = os.path.join(project_root, "resources", "pyensembl_cache")
    
    logging.info(f"Setting pyensembl cache directory to: {pyensembl_cache}")
    os.makedirs(pyensembl_cache, exist_ok=True)
    os.environ['PYENSEBL_CACHE_DIR'] = pyensembl_cache
    
    try:
        ensembl_data = pyensembl.Genome(
            reference_name='GRCh38_local',
            annotation_name='gencode_local',
            gtf_path_or_url=gtf
        )
        ensembl_data.index()
        logging.info("pyensembl initialized successfully.")
    except Exception as e:
        logging.error(f"Failed to initialize pyensembl. Exon-aware search will be skipped. Error: {e}")
        ensembl_data = None

    logging.info("Loading auxiliary data...")
    cancer_genes = set(line.strip() for line in open(cancer_genes_file))
    varscan_variants = set()
    if varscan_vcf:
        logging.info(f"Loading VarScan2 variants from {varscan_vcf}")
        with pysam.VariantFile(varscan_vcf) as vcf:
            for record in vcf: varscan_variants.add(f"{record.chrom}-{record.pos}-{record.ref}-{record.alts[0]}")
    rna_support = {f"{r['contig']}-{r['position']}-{r['refAllele']}-{r['altAllele']}": {'alt_reads': r['altCount']} for _, r in pd.read_csv(rna_counts, sep='\t').iterrows()}

    logging.info("Parsing somatic VCF...")
    somatic_list = []
    with pysam.VariantFile(somatic_vcf) as vcf:
        csq_header = get_csq_header(vcf)
        for record in vcf:
            try: somatic_list.append(parse_vcf_record(record, vcf, csq_header, sample_id))
            except Exception as e: logging.warning(f"Could not parse somatic record {record.chrom}:{record.pos}: {e}")
    somatic_df = pd.DataFrame(somatic_list)
    if somatic_df.empty:
        logging.warning("No somatic variants found. Creating empty report."); Workbook().save(output_xlsx); return

    logging.info("Integrating data and applying somatic tiering...")
    somatic_df["Is Cancer Gene"] = somatic_df["Gene"].isin(cancer_genes)
    somatic_df["VarScan Support"] = somatic_df["VariantKey"].isin(varscan_variants) if varscan_vcf else "-"
    somatic_df["RNA Support (Ref,Alt)"] = somatic_df["VariantKey"].map(lambda k: f"{rna_support.get(k,{}).get('ref_reads',0)},{rna_support.get(k,{}).get('alt_reads',0)}")
    tier_results = somatic_df.apply(lambda r: tier_somatic_variant(r, rna_support.get(r["VariantKey"],{}), r["VarScan Support"], tiering_params), axis=1)
    somatic_df[["Somatic Tier", "Tier Justification"]] = pd.DataFrame(tier_results.tolist(), index=somatic_df.index)

    logging.info("Finding nearby germline variants...")
    nearby_germline_list = []
    if ensembl_data:
        germline_by_chrom = defaultdict(list)
        with pysam.VariantFile(germline_vcf) as vcf:
            for rec in vcf:
                if "missense" in str(rec.info.get("CSQ","")) or "frameshift" in str(rec.info.get("CSQ","")):
                    germline_by_chrom[rec.chrom].append(rec)
        with pysam.VariantFile(germline_vcf) as vcf:
             csq_header = get_csq_header(vcf)
             for _, row in somatic_df.iterrows():
                 for rec in find_nearby_germline_exon_aware(row, germline_by_chrom, ensembl_data):
                     try:
                         parsed = parse_vcf_record(rec, vcf, csq_header, sample_id)
                         parsed["Somatic Tier"] = "Nearby Germline"
                         if parsed not in nearby_germline_list: nearby_germline_list.append(parsed)
                     except Exception as e: logging.warning(f"Could not parse nearby germline record {rec.chrom}:{rec.pos}: {e}")
    
    combined_df = pd.concat([somatic_df, pd.DataFrame(nearby_germline_list)], ignore_index=True) if nearby_germline_list else somatic_df
    combined_df.sort_values(by=["CHROM", "POS"], inplace=True)

    logging.info("Filtering for pathogenic germline variants...")
    pathogenic_germline_list = []
    with pysam.VariantFile(germline_vcf) as vcf:
        csq_header = get_csq_header(vcf)
        for rec in vcf:
             try:
                parsed = parse_vcf_record(rec, vcf, csq_header, sample_id)
                if parsed.get("Impact") == "HIGH" and float(parsed.get('Normal_VAF', 0)) > 0.2:
                    pathogenic_germline_list.append(parsed)
             except Exception as e: logging.warning(f"Could not parse pathogenic germline record {rec.chrom}:{rec.pos}: {e}")
    pathogenic_germline_df = pd.DataFrame(pathogenic_germline_list)
    
    logging.info("Generating Excel report...")
    final_cols = ["CHROM", "POS", "REF", "ALT", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "Impact", "Somatic Tier", "Tumor_DP", "Tumor_AD", "Tumor_VAF", "Normal_DP", "Normal_AD", "Normal_VAF", "RNA Support (Ref,Alt)", "VarScan Support", "Is Cancer Gene", "gnomAD_AF", "COSMIC_CNT", "Tier Justification"]
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Somatic & Nearby Germline Report"
    ws1.append(final_cols)
    header_font = Font(bold=True)
    for cell in ws1[1]: cell.font = header_font
    
    germline_fill = PatternFill("solid", fgColor="FFFF00") # Yellow

    for _, row in combined_df.reindex(columns=final_cols).fillna('/').iterrows():
        ws1.append(list(row))
        if row["Somatic Tier"] == "Nearby Germline":
            for cell in ws1[ws1.max_row]:
                cell.fill = germline_fill
            
    ws2 = wb.create_sheet(title="Pathogenic Germline Variants")
    ws2.append(final_cols)
    for cell in ws2[1]: cell.font = header_font
    if not pathogenic_germline_df.empty:
        for r in dataframe_to_rows(pathogenic_germline_df.reindex(columns=final_cols).fillna('/'), index=False, header=False):
            ws2.append(r)
        
    wb.save(output_xlsx)
    logging.info(f"Report generation complete. Saved to {output_xlsx}")

if __name__ == "__main__":
    main()